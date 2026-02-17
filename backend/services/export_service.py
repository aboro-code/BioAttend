"""
Export service for generating reports
"""

import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from psycopg2.extras import RealDictCursor
from dependencies import get_db_connection


def generate_csv_export():
    """Generate CSV export"""
    conn = get_db_connection()
    df = pd.read_sql(
        """
        SELECT s.name as "Student Name", 
               a.status as "Status", 
               TO_CHAR(a.log_time, 'YYYY-MM-DD HH24:MI:SS') as "Timestamp"
        FROM attendance_logs a 
        JOIN students s ON a.student_id = s.id 
        ORDER BY a.log_time DESC;
    """,
        conn,
    )
    conn.close()

    now = datetime.now()
    filename = f"attendance_{now.strftime('%Y%m%d_%H%M%S')}.csv"

    stream = io.StringIO()
    df.to_csv(stream, index=False)

    return stream.getvalue(), filename


def generate_excel_export():
    """Generate formatted Excel export"""
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        """
        SELECT s.name as student_name, 
               a.status, 
               a.log_time
        FROM attendance_logs a 
        JOIN students s ON a.student_id = s.id 
        ORDER BY a.log_time DESC;
    """
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Styling
    header_fill = PatternFill(
        start_color="4F46E5", end_color="4F46E5", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "BioAttend - Attendance Report"
    title_cell.font = Font(bold=True, size=16, color="4F46E5")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Timestamp
    ws.merge_cells("A2:D2")
    timestamp_cell = ws["A2"]
    timestamp_cell.value = (
        f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    timestamp_cell.font = Font(italic=True, size=10)
    timestamp_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = ["#", "Student Name", "Status", "Timestamp"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Data
    for row_num, record in enumerate(rows, 5):
        ws.cell(row=row_num, column=1, value=row_num - 4).border = border
        ws.cell(row=row_num, column=2, value=record["student_name"]).border = border

        status_cell = ws.cell(row=row_num, column=3, value=record["status"])
        status_cell.border = border
        if record["status"] == "Present":
            status_cell.fill = PatternFill(
                start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"
            )
            status_cell.font = Font(color="065F46", bold=True)

        timestamp_cell = ws.cell(
            row=row_num,
            column=4,
            value=record["log_time"].strftime("%Y-%m-%d %H:%M:%S"),
        )
        timestamp_cell.border = border

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 22

    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    now = datetime.now()
    filename = f"attendance_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

    return excel_file, filename


def generate_detailed_excel_export():
    """Generate detailed Excel with statistics"""
    # Similar to generate_excel_export but with multiple sheets
    # (Copy the logic from your previous excel-detailed endpoint)
    pass  # Implement based on previous code
